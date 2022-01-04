# 方式一
# from django.template import loader # 导入loader方法
# from django.shortcuts import render #导入render 方法
# def test_html(request):
#     t=loader.get_template('test.html')
#     html=t.render({'name':'c语言中文网'})#以字典形式传递数据并生成html
#     return HttpResponse(html) #以 HttpResponse方式响应html
# 方式二
from django.shortcuts import render  # 导入render方法


def test_html(request):
    return render(request, 'test.html', {'name': 'c语言中文网'})  # 根据字典数据生成动态模板


from django.views.decorators.csrf import csrf_exempt
import json
from django.http import HttpResponse, request


@csrf_exempt
def test_getApiResult(request):
    dic = {}
    if request.method == 'GET':
        dic['message'] = 0
        return HttpResponse(json.dumps(dic))
    else:
        dic['message'] = '方法错误'
        return HttpResponse(json.dumps(dic, ensure_ascii=False))


import xlrd


@csrf_exempt
def test1(request):
    # if req.method == 'GET':
    #     return render(req, 'mybook.html',{'book_name':book.name,'icon':icon_url})
    # name = req.POST.get('name')
    # myfile = req.FILES.get('excel')
    # book = Book.objects.create(name=name,icon=myfile)  #将数据存储到Book表中

    # step2 然后读取存入数据库
    new_filePath = "D:\\Users\\O0001673\\Downloads\\test1.xls"
    workbook = xlrd.open_workbook(new_filePath)
    # 获取页签内容，也可以by index获取
    sheetContent = workbook.sheet_by_index(0)
    # sheetContent = workbook.sheet_by_name("安全STR1")
    # 获取总行数和总列数
    # print(sheetContent.nrows, sheetContent.ncols)
    return HttpResponse('OK' + str(sheetContent.nrows) + '' + str(sheetContent.ncols))


@csrf_exempt
def import_case(request):
    """ 导入Excel数据，pk是所属项目的pk """
    if request.method == 'POST':
        try:
            # with transaction.atomic():   # 事物
            # project_pk = request.POST.get("project_pk")  # 数据库使用字段
            excel = request.FILES.get('file_obj')
            book = xlrd.open_workbook(filename=None, file_contents=excel.read())
            sheet = book.sheet_by_index(0)
            title = sheet.row_values(0)

            cell = sheet.cell_value(1, 1)
            # row1 = sheet.nrows[0]
            # value1 = sheet.row_values(row1)

            for row in range(1, sheet.nrows):
                print(sheet.row_values(row))  # 这里取出来每行的数据，就可以写入到数据库了
            return HttpResponse(title)
        except Exception as e:
            print(e)
            return HttpResponse(e)
            # return render(request, 'import_case.html',
            #               {"project_pk": pk, "error": "上传文件类型有误，只支持 xls 和 xlsx 格式的 Excel文档"})
            # return render(request, 'import_case.html', {"project_pk": pk, "error": ""})


from openpyxl import load_workbook


@csrf_exempt
def import_case(request):
    """ 导入Excel数据，pk是所属项目的pk """
    if request.method == 'POST':
        try:
            # with transaction.atomic():   # 事物
            # project_pk = request.POST.get("project_pk")  # 数据库使用字段
            # 加载workbook
            excel = request.FILES.get('file_obj')

            wb = load_workbook(excel)
            sheet = wb['Sheet1']
            cell = sheet.cell(row=1, column=1).value
            cell = sheet.cell(row=1, column=2).value
            cell = sheet.cell(row=1, column=3).value
            print(cell)
            return HttpResponse(cell)
        except Exception as e:
            print(e)
            return HttpResponse(e)
            # return render(request, 'import_case.html',
            #               {"project_pk": pk, "error": "上传文件类型有误，只支持 xls 和 xlsx 格式的 Excel文档"})
            # return render(request, 'import_case.html', {"project_pk": pk, "error": ""})
